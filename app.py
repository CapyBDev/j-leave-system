from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_from_directory
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date
import os
import csv
import io
from datetime import timedelta, datetime
from flask import send_file , Response 
from functools import wraps
from werkzeug.utils import secure_filename
# === REPORT / EXPORT ===
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from weasyprint import HTML, CSS
from uuid import uuid4
from reportlab.lib.pagesizes import A4, landscape


PAGE_SIZE = landscape(A4)
app = Flask(__name__)
app.secret_key = "super_secret_key_change_me"

# === Upload config untuk dokumen sokongan leave ===
LEAVE_UPLOAD_FOLDER = os.path.join("static", "uploads", "leave_docs")
os.makedirs(LEAVE_UPLOAD_FOLDER, exist_ok=True)

ALLOWED_LEAVE_EXTENSIONS = {"png", "jpg", "jpeg", "pdf"}

# === Upload config untuk profile photo ===
PROFILE_UPLOAD_FOLDER = os.path.join("static", "uploads", "profile_photos")
os.makedirs(PROFILE_UPLOAD_FOLDER, exist_ok=True)

# file types allowed for profile photos
ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "gif"}

app.config["PROFILE_UPLOAD_FOLDER"] = PROFILE_UPLOAD_FOLDER


def allowed_leave_file(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_LEAVE_EXTENSIONS
    )

DB_PATH = os.path.join(os.path.dirname(__file__), "database.db")
# ---------------------- Position Hierarchy ----------------------
POSITION_HIERARCHY = {
    "Staff": "Supervisor",
    "Supervisor": "Manager",
    "Manager": "General Manager",
    "General Manager": "CEO",
    "CEO": None  # top of chain
}

def calculate_working_days(start_date, end_date):
    """Return number of working days between start & end, 
    excluding Sat/Sun & public holidays."""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT date FROM holidays")
    holiday_rows = c.fetchall()
    holidays = {h["date"] for h in holiday_rows}
    conn.close()

    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()

    count = 0
    current = start

    while current <= end:
        if current.weekday() < 5:  # 0=Mon ... 4=Fri
            if current.isoformat() not in holidays:
                count += 1
        current += timedelta(days=1)
    return count

def get_next_position(position):
    """Return next higher position for approval or checking chain."""
    return POSITION_HIERARCHY.get(position, None)

def allowed_photo(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# ---------------------- DB Helpers ----------------------
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ✅ AUDIT LOG FUNCTION
def add_log(leave_id, user_id, action, description):
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO leave_logs (leave_id, action, performed_by, timestamp, description)
        VALUES (?,?,?,?,?)
    """, (leave_id, action, user_id, datetime.utcnow().isoformat(), description))
    conn.commit()
    conn.close()


def _add_column_if_missing(cur, table, name, coltype):
    """Idempotent column add (safe for existing DBs)."""
    cur.execute(f"PRAGMA table_info({table})")
    cols = [r["name"] for r in cur.fetchall()]
    if name not in cols:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {name} {coltype}")

def init_db():
    """Create tables if missing and migrate columns if DB already exists."""
    conn = get_db()
    c = conn.cursor()

    # Base tables
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            full_name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin','user')),
            created_at TEXT NOT NULL
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS leaves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            leave_type TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            reason TEXT,
            status TEXT NOT NULL DEFAULT 'Pending',
            created_at TEXT NOT NULL,
            next_approver TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS holidays (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            date TEXT NOT NULL
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )
    """)
    # ✅ Audit log / Timeline table
    c.execute("""
        CREATE TABLE IF NOT EXISTS leave_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            leave_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            performed_by INTEGER NOT NULL,
            timestamp TEXT NOT NULL,
            description TEXT,
            FOREIGN KEY (leave_id) REFERENCES leaves(id),
            FOREIGN KEY (performed_by) REFERENCES users(id)
        )
    """)
    # 🌿 NEW: Leave Applications table (for workflow with checker & approver)
    c.execute("""
        CREATE TABLE IF NOT EXISTS leave_applications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            full_name TEXT,
            position TEXT,
            leave_type TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            total_days INTEGER,
            reason TEXT,
            status TEXT NOT NULL DEFAULT 'Pending Check',
            checker_name TEXT,
            approver_name TEXT,
            support_doc TEXT,
            contact_address TEXT,
            contact_phone TEXT,
            created_at TEXT NOT NULL,
            checked_at TEXT,
            approved_at TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
        # NEW: MC records table (medical certificates)
    c.execute("""
        CREATE TABLE IF NOT EXISTS mc_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            mc_number TEXT,
            start_date TEXT,
            end_date TEXT,
            pdf_path TEXT,
            uploaded_by INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (uploaded_by) REFERENCES users(id)
        )
    """)


    # Migrations (safe if already applied)
    _add_column_if_missing(c, "users", "entitlement", "INTEGER DEFAULT 0")
    _add_column_if_missing(c, "users", "department_id", "INTEGER REFERENCES departments(id)")
    _add_column_if_missing(c, "users", "position", "TEXT")
    _add_column_if_missing(c, "users", "approver_role", "TEXT")
    
    conn.commit()

    conn = get_db()
    c = conn.cursor()

    # Safely add new columns if they don't exist yet (for users table)
    def add_col_if_not_exists(colname, coldef):
        try:
            c.execute(f"ALTER TABLE users ADD COLUMN {colname} {coldef}")
        except:
            pass  # column already exists

    add_col_if_not_exists("ic_number", "TEXT")
    add_col_if_not_exists("email", "TEXT")
    add_col_if_not_exists("phone", "TEXT")
    add_col_if_not_exists("address", "TEXT")
    add_col_if_not_exists("enrollment_date", "TEXT")
    add_col_if_not_exists("availability", "TEXT DEFAULT 'Available'")
    add_col_if_not_exists("profile_photo", "TEXT")
    add_col_if_not_exists("reset_token", "TEXT")
    add_col_if_not_exists("reset_token_expiry", "TEXT")


    # 🌿 NEW: Safely add columns to the leaves table
    def add_leaves_col_if_not_exists(colname, coldef):
        try:
            c.execute(f"ALTER TABLE leaves ADD COLUMN {colname} {coldef}")
        except:
            pass  # column already exists

    add_leaves_col_if_not_exists("contact_address", "TEXT")
    add_leaves_col_if_not_exists("contact_phone", "TEXT")
    add_leaves_col_if_not_exists("notes", "TEXT")
    # add_leaves_col_if_not_exists("next_approver", "TEXT")
    add_leaves_col_if_not_exists("checked_by_position", "TEXT")
    add_leaves_col_if_not_exists("checked_status", "TEXT DEFAULT 'Pending'")
    add_leaves_col_if_not_exists("next_approver_position", "TEXT")
    add_leaves_col_if_not_exists("next_approver_department", "TEXT")
    add_leaves_col_if_not_exists("checked_by_user_id", "INTEGER")
    add_leaves_col_if_not_exists("approved_by_user_id", "INTEGER")

    conn.commit()

    # Seed admin + sample user
    c.execute("SELECT 1 FROM users WHERE username=?", ('admin',))
    if not c.fetchone():
        c.execute(
            "INSERT INTO users (username, full_name, password_hash, role, created_at, entitlement) VALUES (?,?,?,?,?,?)",
            ("admin", "Administrator", generate_password_hash("admin123"), "admin",
             datetime.utcnow().isoformat(), 20)
        )

    c.execute("SELECT 1 FROM users WHERE username=?", ('alice',))
    if not c.fetchone():
        c.execute(
            "INSERT INTO users (username, full_name, password_hash, role, created_at, entitlement) VALUES (?,?,?,?,?,?)",
            ("alice", "Alice Johari", generate_password_hash("password"), "user",
             datetime.utcnow().isoformat(), 14)
        )

    # Seed simple holidays (only if empty)
    c.execute("SELECT COUNT(*) AS cnt FROM holidays")
    if c.fetchone()["cnt"] == 0:
        holidays = [
            ("New Year's Day", f"{date.today().year}-01-01"),
            ("Labour Day", f"{date.today().year}-05-01"),
            ("Malaysia Day", f"{date.today().year}-09-16"),
        ]
        c.executemany("INSERT INTO holidays (name, date) VALUES (?,?)", holidays)

    conn.commit()
    conn.close()


@app.before_request
def ensure_db():
    # Always run init_db to ensure migrations apply to older DBs
    init_db()
def auto_reset_mc_availability():
    today = date.today().isoformat()
    conn = get_db()
    c = conn.cursor()

    # Cari user yang MC sudah tamat tapi masih status MC
    c.execute("""
        SELECT DISTINCT u.id
        FROM mc_records m
        JOIN users u ON u.id = m.user_id
        WHERE m.end_date IS NOT NULL
        AND date(m.end_date) < date(?)
        AND u.availability = 'MC'
    """, (today,))

    users_to_reset = c.fetchall()

    for u in users_to_reset:
        c.execute(
            "UPDATE users SET availability='Available' WHERE id=?",
            (u["id"],)
        )

    conn.commit()
    conn.close()

@app.before_request
def before_any_request():
    auto_reset_mc_availability()

# ---------------------- Auth ----------------------
@app.route("/")
def home():
    if "user_id" in session:
        return redirect(url_for("admin_dashboard" if session.get("role") == "admin" else "user_dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    # Initialise counter
    if "login_attempts" not in session:
        session["login_attempts"] = 0

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        # If too many attempts, just show message + forgot link
        if session["login_attempts"] >= 3:
            flash("Too many failed attempts. Please use 'Forgot password / username'.", "warning")
            return redirect(url_for("login"))

        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE username=?", (username,))
        user = c.fetchone()
        conn.close()

        if user and check_password_hash(user["password_hash"], password):
            # ✅ success
            session["login_attempts"] = 0
            session.update({
                "user_id": user["id"],
                "username": user["username"],
                "full_name": user["full_name"],
                "role": user["role"],
                "position": user["position"],
                "profile_photo": user["profile_photo"]
            })

            flash("Login success. Welcome back!", "success")
            return redirect(url_for("admin_dashboard" if user["role"] == "admin" else "user_dashboard"))

        # ❌ wrong
        session["login_attempts"] += 1
        flash("Invalid username or password.", "danger")

    # show_forgot = True bila dah 3 kali salah
    return render_template("login.html", show_forgot=session.get("login_attempts", 0) >= 3)

@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))

# ---------------------- Decorators ----------------------
def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Admin access only.", "warning")
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

@app.route("/admin/mc/update/<int:mc_id>", methods=["POST"])
@admin_required
def admin_update_mc(mc_id):
    mc_number = request.form.get("mc_number")
    start = request.form.get("start_date")
    end = request.form.get("end_date")

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        UPDATE mc_records
        SET mc_number=?, start_date=?, end_date=?
        WHERE id=?
    """, (mc_number, start, end, mc_id))
    conn.commit()
    conn.close()

    flash("MC updated.", "success")
    return redirect(url_for("admin_dashboard"))

def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            flash("Please login first.", "warning")
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

@app.route("/admin/leaves/<view>")
def admin_leave_modal(view):
    if view == "today":
        data = get_today_leaves() 
    elif view == "pending":
        data = get_pending_leaves()
    else:
        data = get_all_leaves()

    return render_template("admin/modal_table.html", data=data)

@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    # ambil filter tarikh dari query string
    date_from = request.args.get("date_from", "").strip()
    date_to   = request.args.get("date_to", "").strip()

    # pass to dashboard calculation
    data = get_dashboard_data(
        date_from=date_from or None,
        date_to=date_to or None
    )

    # untuk pre-fill input form
    data["date_from"] = date_from
    data["date_to"]   = date_to

    return render_template("admin_dashboard.html", **data)

def get_leave_data(mode):
    conn = get_db()
    c = conn.cursor()
    if mode == "weekly":
        query = """SELECT u.full_name, l.leave_type, l.start_date, l.end_date, l.status
                   FROM leaves l JOIN users u ON l.user_id=u.id
                   WHERE date(l.start_date)>=date('now','-7 day')
                   ORDER BY l.start_date ASC"""
    else:
        query = """SELECT u.full_name, l.leave_type, l.start_date, l.end_date, l.status
                   FROM leaves l JOIN users u ON l.user_id=u.id
                   ORDER BY u.full_name ASC"""
    c.execute(query)
    data = c.fetchall()
    conn.close()
    return data

@app.route("/export/leave/pdf")
def export_pdf():
    mode = request.args.get("mode","weekly")
    data = get_leave_data(mode)
    stream = io.BytesIO()
    pdf = canvas.Canvas(stream, pagesize=A4)
    pdf.setFont("Helvetica-Bold",16); pdf.drawString(40,800,f"Leave Report ({mode.title()})")
    pdf.setFont("Helvetica",10); y=770
    pdf.drawString(40,y,"Name"); pdf.drawString(180,y,"Type")
    pdf.drawString(260,y,"Start"); pdf.drawString(340,y,"End"); pdf.drawString(420,y,"Status"); y-=20
    for r in data:
        pdf.drawString(40,y,r["full_name"])
        pdf.drawString(180,y,r["leave_type"])
        pdf.drawString(260,y,r["start_date"])
        pdf.drawString(340,y,r["end_date"])
        pdf.drawString(420,y,r["status"])
        y -= 20
    pdf.save(); stream.seek(0)
    return send_file(stream, as_attachment=True,
                     download_name=f"LeaveReport_{mode}.pdf")
    

    
def export_leave_excel(data, year):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    headers = ["ID","Name","Leave Type"] + MONTHS + ["Total Used","Remaining"]
    ws.append(headers)

    for r in data:
        row = [
            r["user_id"],
            r["name"],
            r["leave_type"],
            *[r["months"][m] for m in MONTHS],
            r["total_used"],
            r["remaining"]
        ]
        ws.append(row)

    file = io.BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file,
        as_attachment=True,
        download_name=f"Leave_Report_{year}.xlsx")


@app.route("/admin/leave-report/employee/<int:user_id>/download")
@admin_required
def download_employee_leave_report(user_id):
    month = request.args.get("month")
    year = request.args.get("year")

    conn = get_db()
    c = conn.cursor()

    query = """
        SELECT leave_type, start_date, end_date, total_days
        FROM leave_applications
        WHERE user_id=? AND status='Approved'
    """
    params = [user_id]

    if month:
        query += " AND strftime('%Y-%m', start_date)=?"
        params.append(month)
    elif year:
        query += " AND strftime('%Y', start_date)=?"
        params.append(year)

    c.execute(query, params)
    rows = c.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Leave Type", "Start Date", "End Date", "Days"])
    for r in rows:
        writer.writerow([r["leave_type"], r["start_date"], r["end_date"], r["total_days"]])

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=employee_{user_id}_leave_report.csv"}
    )

from datetime import timedelta

def daterange(start, end):
    for n in range((end - start).days + 1):
        yield start + timedelta(days=n)

@app.route("/leave-report/individual/<int:user_id>")
def individual_leave_preview(user_id):
    year = request.args.get("year", datetime.now().year)

    report = build_individual_leave_report(user_id, year)

    return render_template(
        "reports/individual_leave_preview.html",
        employee=report["employee"],
        department=report["department"],
        monthly=report["monthly"],
        summary=report["summary"],
        year=year
    )

@app.route("/leave-report/individual/<int:user_id>/download")
def download_individual_leave_pdf(user_id):
    year = request.args.get("year", datetime.now().year)

    report = build_individual_leave_report(user_id, year)

    html = render_template(
        "reports/individual_leave_report.html",
        report=report,
        year=year
    )

    pdf = HTML(string=html).write_pdf(
        stylesheets=[CSS(string="""
            body { font-family: Arial; font-size: 12px; }
        """)]
    )

    return Response(
        pdf,
        mimetype="application/pdf",
        headers={
            "Content-Disposition":
            f"attachment; filename=leave_report_{report['employee']['name']}.pdf"
        }
    )

@app.route("/leave-report/<int:user_id>/<string:format>")
def download_individual_leave_report(user_id, format):
    conn = sqlite3.connect("database.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT
            u.full_name,
            d.name AS department_name,
            l.leave_type,
            l.start_date,
            l.end_date
        FROM leaves l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.id = ? AND l.status = 'Approved'
        ORDER BY l.start_date
    """, (user_id,))
    rows = cur.fetchall()
    conn.close()

    if format == "excel":
        import pandas as pd
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        return send_file(
            output,
            download_name="leave_report.xlsx",
            as_attachment=True
        )

    # ===== PDF =====
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)

    y = 800
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, f"Leave Report - {rows[0]['full_name']}")
    y -= 30

    c.setFont("Helvetica", 10)
    for r in rows:
        c.drawString(
            50, y,
            f"{r['start_date']} → {r['end_date']} | {r['leave_type']}"
        )
        y -= 18
        if y < 50:
            c.showPage()
            y = 800

    c.save()
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="leave_report.pdf")

@app.route("/admin/dashboard/view/<filter>")
@admin_required
def admin_dashboard_view(filter):
    conn = get_db()
    c = conn.cursor()

    # ====== Ambil nilai filter dari query string ======
    filter_type     = request.args.get("type", "").strip()
    filter_position = request.args.get("position", "").strip()
    filter_alpha    = request.args.get("alpha", "").strip()     # A, B, C...
    date_from       = request.args.get("date_from", "").strip()
    date_to         = request.args.get("date_to", "").strip()

    # ====== Base query ikut "filter" dalam URL (today/month/pending...) ======
    base_query = "SELECT * FROM leave_applications WHERE 1=1 "
    params = []

    if filter == "month":
        title = "Leave Applications This Month"
        base_query += " AND strftime('%m', start_date) = strftime('%m','now') "

    elif filter == "today":
        title = "Employees On Leave Today"
        base_query += """
            AND status='Approved'
            AND date('now') BETWEEN start_date AND end_date
        """

    elif filter == "pending":
        title = "Pending Leave Requests"
        base_query += " AND status IN ('Pending Check','Pending Approval') "

    # 🔹 TAMBAH BAHAGIAN INI
    elif filter == "approved":
        title = "Approved Leave Requests"
        base_query += " AND status='Approved' "

    elif filter == "rejected":
        title = "Rejected Leave Requests"
        base_query += " AND status='Rejected' "

    else:
        title = "Leave Details"

    # ====== Tambah FILTER TYPE (leave_type) ======
    if filter_type:
        base_query += " AND leave_type = ? "
        params.append(filter_type)

    # ====== Tambah FILTER POSITION ======
    if filter_position:
        base_query += " AND position = ? "
        params.append(filter_position)

    # ====== Tambah FILTER ALPHABET NAMA (full_name bermula dengan huruf) ======
    if filter_alpha:
        base_query += " AND full_name LIKE ? "
        params.append(f"{filter_alpha}%")

    # ====== Tambah FILTER TARIKH (range start_date) ======
    # date_from & date_to format: YYYY-MM-DD (HTML input type="date")
    if date_from and date_to:
        base_query += " AND date(start_date) BETWEEN date(?) AND date(?) "
        params.extend([date_from, date_to])
    elif date_from:
        base_query += " AND date(start_date) >= date(?) "
        params.append(date_from)
    elif date_to:
        base_query += " AND date(start_date) <= date(?) "
        params.append(date_to)

    # Susun ikut tarikh
    base_query += " ORDER BY date(start_date) DESC "

    c.execute(base_query, params)
    rows = c.fetchall()

    # ====== Data untuk dropdown filter (type & position) ======
    c.execute("SELECT DISTINCT leave_type FROM leave_applications WHERE leave_type IS NOT NULL")
    leave_types = [r[0] for r in c.fetchall()]

    c.execute("SELECT DISTINCT position FROM leave_applications WHERE position IS NOT NULL")
    positions = [r[0] for r in c.fetchall()]

    conn.close()

    return render_template(
        "admin_dashboard_detail.html",
        title=title,
        leaves=rows,
        leave_types=leave_types,
        positions=positions,
        # untuk isi semula value dalam form
        filter_type=filter_type,
        filter_position=filter_position,
        filter_alpha=filter_alpha,
        date_from=date_from,
        date_to=date_to,
        current_filter=filter,
    )
@app.route("/admin/leave-report/view")
@admin_required
def view_all_leave_report():
    month = request.args.get("month")
    year = request.args.get("year")

    conn = get_db()
    c = conn.cursor()

    query = """
        SELECT
            la.full_name,
            d.name AS department,
            COUNT(*) AS total_applications,
            SUM(la.total_days) AS total_days
        FROM leave_applications la
        LEFT JOIN users u ON la.user_id=u.id
        LEFT JOIN departments d ON u.department_id=d.id
        WHERE la.status='Approved'
    """
    params = []

    if month:
        query += " AND strftime('%Y-%m', la.start_date)=?"
        params.append(month)
    elif year:
        query += " AND strftime('%Y', la.start_date)=?"
        params.append(year)

    query += " GROUP BY la.user_id ORDER BY la.full_name"

    c.execute(query, params)
    rows = c.fetchall()
    conn.close()

    return render_template(
        "reports/all_leave_report.html",
        rows=rows,
        month=month,
        year=year
    )

def get_dashboard_data(date_from=None, date_to=None):
    conn = get_db()
    c = conn.cursor()

    # ===================== SUMMARY CARDS =======================
    # Total leave application this month (submitted)
    c.execute("""
        SELECT COUNT(*) 
        FROM leave_applications 
        WHERE strftime('%m', start_date) = strftime('%m','now')
    """)
    total_this_month = c.fetchone()[0]

    # Total employees on leave today
    c.execute("""
        SELECT COUNT(*) 
        FROM leave_applications
        WHERE status = 'Approved'
        AND date('now') BETWEEN start_date AND end_date
    """)
    leave_today = c.fetchone()[0]

    # Pending leave (check + approval stage)
    c.execute("""
        SELECT COUNT(*)
        FROM leave_applications
        WHERE status IN ('Pending Check','Pending Approval')
    """)
    pending_leave = c.fetchone()[0]

    # Rejected leave
    c.execute("""
        SELECT COUNT(*)
        FROM leave_applications
        WHERE status = 'Rejected'
    """)
    rejected_leave = c.fetchone()[0]

    # ===================== RECENT REQUEST LIST ==================
    c.execute("""
        SELECT id, full_name, leave_type, status
        FROM leave_applications
        ORDER BY created_at DESC LIMIT 7
    """)
    recent_requests = c.fetchall()

    # ===================== ON LEAVE TODAY LIST ==================
    c.execute("""
        SELECT 
            la.full_name, 
            COALESCE(d.name, '-') AS department, 
            la.end_date AS return_date
        FROM leave_applications la
        LEFT JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status = 'Approved'
        AND date('now') BETWEEN la.start_date AND la.end_date
    """)
    on_leave_today = c.fetchall()

    # ===================== DEPARTMENTS FOR FILTER ================
    c.execute("SELECT DISTINCT name FROM departments WHERE name IS NOT NULL")
    departments = [row[0] for row in c.fetchall()]

    # ===================== PIE (LEAVE TYPES) =====================
    c.execute("""
        SELECT leave_type, COUNT(*) 
        FROM leave_applications 
        GROUP BY leave_type
    """)
    rows = c.fetchall()
    leave_types = [r[0] for r in rows]
    leave_type_count = [r[1] for r in rows]

    # ===================== TREND (WEEKLY) ========================
    c.execute("""
        SELECT strftime('%d/%m', start_date), COUNT(*) 
        FROM leave_applications
        WHERE date(start_date) >= date('now','-7 day')
        GROUP BY start_date
    """)
    t = c.fetchall()
    trend_labels = [r[0] for r in t]
    trend_data = [r[1] for r in t]
    
    # ===================== MC RECORDS =====================
    c.execute("""
        SELECT 
            m.id,
            u.full_name,
            u.address,
            m.mc_number,
            m.start_date,
            m.end_date,
            m.pdf_path,
            m.created_at
        FROM mc_records m
        JOIN users u ON u.id = m.user_id
        ORDER BY m.created_at DESC
        LIMIT 10
    """)
    mc_records = c.fetchall()
    # ===================== MC COUNT PER MONTH =====================
    c.execute("""
        SELECT strftime('%m/%Y', created_at) AS month, COUNT(*)
        FROM mc_records
        GROUP BY month
        ORDER BY created_at DESC
        LIMIT 6
    """)
    mc_rows = c.fetchall()

    mc_labels = [r[0] for r in mc_rows]
    mc_counts = [r[1] for r in mc_rows]

    # ===================== APPROVED / REJECTED DETAILS ===========
    # range tarikh untuk dashboard (filter ikut start_date)
    range_clauses = []
    if date_from:
        range_clauses.append("date(start_date) >= date(?)")
    if date_to:
        range_clauses.append("date(start_date) <= date(?)")

    where_range = ""
    if range_clauses:
        where_range = " AND " + " AND ".join(range_clauses)

    # --- Rejected details ---
    params_rej = []
    if date_from:
        params_rej.append(date_from)
    if date_to:
        params_rej.append(date_to)

    c.execute("""
        SELECT id, full_name, leave_type, start_date, end_date, status
        FROM leave_applications
        WHERE status = 'Rejected' """ + where_range + """
        ORDER BY date(start_date) DESC
        LIMIT 10
    """, params_rej)
    rejected_details = c.fetchall()

    # --- Approved details ---
    params_app = []
    if date_from:
        params_app.append(date_from)
    if date_to:
        params_app.append(date_to)

    c.execute("""
        SELECT id, full_name, leave_type, start_date, end_date, status
        FROM leave_applications
        WHERE status = 'Approved' """ + where_range + """
        ORDER BY date(start_date) DESC
        LIMIT 10
    """, params_app)
    approved_details = c.fetchall()

    conn.close()

    return {
        "total_this_month": total_this_month,
        "leave_today": leave_today,
        "pending_leave": pending_leave,
        "rejected_leave": rejected_leave,
        "recent_requests": recent_requests,
        "on_leave_today": on_leave_today,
        "departments": departments,
        "leave_types": leave_types,
        "leave_type_count": leave_type_count,
        "trend_labels": trend_labels,
        "trend_data": trend_data,
        "inc_month": 12,
        "diff_today": 3,
        "rejected_details": rejected_details,
        "approved_details": approved_details,
        "mc_records": mc_records,
        "mc_labels": mc_labels,
        "mc_counts": mc_counts,

    }


def get_leave_data(mode):
    conn = get_db()
    c = conn.cursor()

    if mode == "weekly":
        query = """
            SELECT u.full_name, l.leave_type, l.start_date, l.end_date, l.status
            FROM leaves l
            JOIN users u ON l.user_id = u.id
            WHERE date(l.start_date) >= date('now','-7 day')
            ORDER BY l.start_date ASC
        """
    else:  # alphabetical
        query = """
            SELECT u.full_name, l.leave_type, l.start_date, l.end_date, l.status
            FROM leaves l
            JOIN users u ON l.user_id = u.id
            ORDER BY u.full_name ASC
        """

    c.execute(query)
    rows = c.fetchall()
    conn.close()
    return rows

@app.route("/admin/users")
@admin_required
def manage_users():
    conn = get_db()
    c = conn.cursor()

    today = date.today().isoformat()

    filter_name  = request.args.get("name", "")
    filter_dept  = request.args.get("department", "")
    filter_avail = request.args.get("availability", "")
    sort         = request.args.get("sort", "id_desc")

    base_query = """
        SELECT 
            u.id,
            u.username,
            u.full_name,
            u.role,
            u.position,
            u.entitlement,
            u.enrollment_date,
            u.created_at,
            u.department_id,
            u.ic_number,
            u.email,
            u.phone,
            u.address,

            CASE
                WHEN EXISTS (
                    SELECT 1 FROM mc_records m
                    WHERE m.user_id = u.id
                    AND m.start_date IS NOT NULL
                    AND m.end_date IS NOT NULL
                    AND date(?) BETWEEN date(m.start_date) AND date(m.end_date)
                ) THEN 'MC'
                WHEN EXISTS (
                    SELECT 1 FROM leave_applications la
                    WHERE la.user_id = u.id
                    AND la.status = 'Approved'
                    AND date(?) BETWEEN date(la.start_date) AND date(la.end_date)
                ) THEN 'On Leave'
                ELSE 'Available'
            END AS availability,

            d.name AS department_name

        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE 1=1
    """
    params = [today, today]


    if filter_name:
        base_query += " AND u.full_name LIKE ?"
        params.append(f"%{filter_name}%")

    if filter_dept:
        base_query += " AND d.name = ?"
        params.append(filter_dept)

    if filter_avail:
        base_query += """
            AND (
                CASE
                    WHEN EXISTS (
                        SELECT 1 FROM leave_applications la
                        WHERE la.user_id = u.id
                        AND la.status = 'Approved'
                        AND date(?) BETWEEN la.start_date AND la.end_date
                    ) THEN 'On Leave'
                    ELSE COALESCE(u.availability, 'Available')
                END
            ) = ?
        """
        params.extend([today, filter_avail])

    if sort == "az":
        base_query += " ORDER BY u.full_name ASC"
    elif sort == "za":
        base_query += " ORDER BY u.full_name DESC"
    else:
        base_query += " ORDER BY u.id DESC"

    c.execute(base_query, params)
    users = [dict(row) for row in c.fetchall()]

    c.execute("SELECT * FROM departments ORDER BY name")
    departments = [dict(row) for row in c.fetchall()]

    conn.close()

    return render_template(
        "manage_users.html",
        users=users,
        departments=departments,
        current_user_role=session.get("role")
    )

@app.route("/admin/users/update_availability/<int:user_id>", methods=["POST"])
@admin_required
def update_availability(user_id):
    """
    AJAX endpoint:
    Expect JSON body: { "availability": "Available" | "Out" | "MC" | "WFH" }
    Returns JSON { success: True, prev: "<old>" } or error.
    """
    import json
    try:
        data = request.get_json(force=True)
    except Exception:
        return jsonify({"success": False, "error": "Invalid JSON"}), 400

    new_status = (data.get("availability") or "").strip()
    if new_status not in ("Available", "Out", "MC", "WFH"):
        return jsonify({"success": False, "error": "Invalid availability"}), 400

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT availability FROM users WHERE id=?", (user_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "error": "User not found"}), 404

    prev = row["availability"] if row["availability"] is not None else ""

    try:
        c.execute("UPDATE users SET availability=? WHERE id=?", (new_status, user_id))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"success": False, "error": f"DB error: {str(e)}"}), 500

    conn.close()
    return jsonify({"success": True, "prev": prev})

@app.route("/admin/upload_mc", methods=["POST"])
@admin_required
def upload_mc():
    """
    Form fields expected:
      - user_id
      - mc_number (optional)
      - mc_start (optional, YYYY-MM-DD)
      - mc_end (optional)
      - mc_pdf (file, required)
    After saving file -> insert into mc_records, set users.availability='Out'
    """
    user_id = request.form.get("user_id")
    mc_number = request.form.get("mc_number", "").strip()
    mc_start = request.form.get("mc_start", "").strip() or None
    mc_end = request.form.get("mc_end", "").strip() or None
    pdf = request.files.get("mc_pdf")

    if not user_id:
        flash("Sila pilih pengguna untuk MC.", "warning")
        return redirect(url_for("manage_users"))

    # Validate user exists
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM users WHERE id=?", (user_id,))
    if not c.fetchone():
        conn.close()
        flash("Pengguna tidak ditemui.", "danger")
        return redirect(url_for("manage_users"))

    if not pdf or pdf.filename == "":
        conn.close()
        flash("Sila muat naik fail MC (PDF/JPG/PNG).", "warning")
        return redirect(url_for("manage_users"))

    # Validate extension
    filename = secure_filename(pdf.filename)
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in ALLOWED_LEAVE_EXTENSIONS:
        conn.close()
        flash("Jenis fail tidak dibenarkan. Gunakan PDF/PNG/JPG.", "danger")
        return redirect(url_for("manage_users"))

    # Save file
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    save_name = f"mc_{user_id}_{timestamp}_{filename}"
    save_path = os.path.join(LEAVE_UPLOAD_FOLDER, save_name)
    try:
        pdf.save(save_path)
    except Exception as e:
        conn.close()
        flash("Gagal simpan fail: " + str(e), "danger")
        return redirect(url_for("manage_users"))

    # Insert MC record + set availability to Out
    try:
        c.execute("""
            INSERT INTO mc_records (user_id, mc_number, start_date, end_date, pdf_path, uploaded_by, created_at)
            VALUES (?,?,?,?,?,?,?)
        """, (user_id, mc_number, mc_start, mc_end, save_name, session.get("user_id"), datetime.utcnow().isoformat()))

        c.execute("UPDATE users SET availability=? WHERE id=?", ("MC", user_id))
        conn.commit()
        flash("MC berjaya dimuat naik dan status pengguna dikemaskini.", "success")
    except Exception as e:
        conn.rollback()
        flash("Gagal simpan rekod MC: " + str(e), "danger")
    finally:
        conn.close()

    return redirect(url_for("manage_users"))

@app.route("/api/mc-trend")
@admin_required
def mc_trend_api():
    view = request.args.get("view", "monthly")  # weekly | monthly

    conn = get_db()
    c = conn.cursor()

    if view == "weekly":
        # 🗓️ Last 7 days (including today)
        c.execute("""
            SELECT 
                strftime('%d/%m', created_at) AS label,
                COUNT(*) AS total
            FROM mc_records
            WHERE date(created_at) >= date('now','-6 day')
            GROUP BY date(created_at)
            ORDER BY date(created_at)
        """)
    else:
        # 🗓️ Last 6 months
        c.execute("""
            SELECT 
                strftime('%m/%Y', created_at) AS label,
                COUNT(*) AS total
            FROM mc_records
            WHERE date(created_at) >= date('now','-5 months')
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY strftime('%Y-%m', created_at)
        """)

    rows = c.fetchall()
    conn.close()

    labels = [r["label"] for r in rows]
    data   = [r["total"] for r in rows]

    return jsonify({
        "labels": labels,
        "data": data
    })


@app.route("/admin/users/upload_photo/<int:user_id>", methods=["POST"])
def upload_profile_photo(user_id):
    file = request.files.get("photo")

    if not file:
        return "No file received", 400

    filename = f"user_{user_id}.jpg"
    filepath = os.path.join("static/uploads/profile/", filename)
    file.save(filepath)

    db.session.query(User).filter(User.id == user_id).update({ "profile_photo": filename })
    db.session.commit()

    return "OK", 200


@app.route("/api/user/<int:user_id>")
@admin_required
def api_user(user_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT 
            u.*,
            d.name AS department_name
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.id = ?
    """, (user_id,))
    user = c.fetchone()
    conn.close()

    if not user:
        return jsonify({"error": "User not found"}), 404

    return jsonify(dict(user))

@app.route("/admin/users/create", methods=["POST"])
@admin_required
def create_user():
    username        = request.form.get("username", "").strip()
    full_name       = request.form.get("full_name", "").strip()
    role            = request.form.get("role", "user")
    password        = request.form.get("password", "")
    department_id   = request.form.get("dept_id") or None
    position        = request.form.get("position", "").strip()
    enrollment_date = request.form.get("enrollment_date", "").strip()
    entitlement     = request.form.get("entitlement", "").strip()

    # extra fields
    email      = request.form.get("email", "").strip()
    phone      = request.form.get("phone", "").strip()
    ic_number  = request.form.get("ic_number", "").strip()
    address    = request.form.get("address", "").strip()

    if not username or not full_name or not password:
        flash("All fields are required.", "danger")
        return redirect(url_for("manage_users"))

    try:
        entitlement_val = int(entitlement) if entitlement != "" else 0
    except ValueError:
        entitlement_val = 0

    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""
            INSERT INTO users (
                username,
                full_name,
                password_hash,
                role,
                created_at,
                department_id,
                position,
                enrollment_date,
                entitlement,
                email,
                phone,
                ic_number,
                address
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            username,
            full_name,
            generate_password_hash(password),
            role,
            datetime.utcnow().isoformat(),
            department_id,
            position,
            enrollment_date,
            entitlement_val,
            email,
            phone,
            ic_number,
            address
        ))
        conn.commit()
        flash("User created successfully.", "success")
    except sqlite3.IntegrityError:
        flash("Username already exists.", "danger")
    finally:
        conn.close()
    return redirect(url_for("manage_users"))


@app.route("/admin/users/delete/<int:user_id>", methods=["POST"])
@admin_required
def delete_user(user_id):
    if user_id == session.get("user_id"):
        flash("You cannot delete yourself.", "warning")
        return redirect(url_for("manage_users"))
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    flash("User deleted.", "info")
    return redirect(url_for("manage_users"))

@app.route("/admin/users/update/<int:user_id>", methods=["POST"])
@admin_required
def update_user_details(user_id):
    conn = get_db()
    c = conn.cursor()

    # get entitlement safely (default 0)
    entitlement = request.form.get("entitlement")
    entitlement = int(entitlement) if entitlement and entitlement.isdigit() else 0

    c.execute("""
        UPDATE users SET 
            full_name     = ?,
            email         = ?,
            phone         = ?,
            address       = ?,
            position      = ?,
            department_id = ?,
            availability  = ?,
            entitlement   = ?
        WHERE id = ?
    """, (
        request.form.get("full_name"),
        request.form.get("email"),
        request.form.get("phone"),
        request.form.get("address"),
        request.form.get("position"),
        request.form.get("dept_id") or None,
        request.form.get("availability"),
        entitlement,
        user_id
    ))

    conn.commit()
    conn.close()

    return "", 204

@app.route("/admin/users/entitlement/<int:user_id>", methods=["POST"])
@admin_required
def update_entitlement(user_id):
    """<<< This fixes the missing endpoint used by the template >>>"""
    raw = request.form.get("entitlement", "0").strip()
    try:
        value = max(0, int(raw))
    except ValueError:
        flash("Entitlement must be a whole number.", "danger")
        return redirect(url_for("manage_users"))

    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE users SET entitlement=? WHERE id=?", (value, user_id))
    conn.commit()
    conn.close()
    flash("Entitlement updated.", "success")
    return redirect(url_for("manage_users"))

@app.route("/manage_leaves")
def manage_leaves():
    from datetime import datetime, timedelta

    conn = get_db()
    cur = conn.cursor()

    # =========================================================
    # QUERY PARAMS
    # =========================================================

    # ---- Leave Report (sticky state) ----
    report_year = request.args.get("year", datetime.now().strftime("%Y"))
    selected_department = request.args.get("department", "all")
    action = request.args.get("action")
    should_build_report = (
        action in ("filter", "search") or request.args.get("year")
    )
    
    # ---- Monthly Matrix ----
    selected_month = request.args.get(
        "matrix_month", datetime.now().strftime("%m")
    )
    matrix_year = request.args.get(
        "matrix_year", datetime.now().strftime("%Y")
    )
    selected_dept = request.args.get("matrix_department", "all")

    view = request.args.get("view", "monthly_matrix")

    # =========================================================
    # DEPARTMENTS
    # =========================================================
    cur.execute("SELECT id, name FROM departments ORDER BY name")
    departments = cur.fetchall()

    # =========================================================
    # YEARLY LEAVE REPORT (FIXED – ACCURATE DAYS)
    # =========================================================
    leave_report = []

    MONTH_MAP = {
        "01": "JAN", "02": "FEB", "03": "MAR", "04": "APR",
        "05": "MAY", "06": "JUN", "07": "JUL", "08": "AUG",
        "09": "SEP", "10": "OCT", "11": "NOV", "12": "DEC"
    }

    if should_build_report:
        params = [report_year]
        dept_filter = ""

        if selected_department != "all":
            dept_filter = "AND d.name = ?"
            params.append(selected_department)

        # 🔥 IMPORTANT: NO GROUP BY
        cur.execute(f"""
            SELECT
                u.id AS user_id,
                u.full_name,
                la.leave_type,
                la.start_date,
                la.end_date,
                la.total_days,
                u.entitlement
            FROM leave_applications la
            JOIN users u ON u.id = la.user_id
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE la.status = 'Approved'
            AND strftime('%Y', la.start_date) = ?
            {dept_filter}
            ORDER BY u.full_name, la.start_date
        """, params)

        rows = cur.fetchall()
        users = {}

        for r in rows:
            uid = r["user_id"]

            if uid not in users:
                users[uid] = {
                    "user_id": uid,
                    "name": r["full_name"],
                    "monthly": {m: 0 for m in MONTH_MAP.values()},
                    "monthly_details": {},
                    "leave_type_details": {},
                    "total_used": 0,
                    "remaining": r["entitlement"] or 0
                }

            if r["leave_type"] == "MC":
                continue

            start = datetime.strptime(r["start_date"], "%Y-%m-%d")
            end   = datetime.strptime(r["end_date"], "%Y-%m-%d")

            cur_day = start
            while cur_day <= end:
                month_key = MONTH_MAP[cur_day.strftime("%m")]

                # ✔️ Count working days only
                if cur_day.weekday() < 5:
                    users[uid]["monthly"][month_key] += 1
                    users[uid]["total_used"] += 1
                    users[uid]["remaining"] -= 1

                    # ---- monthly details ----
                    users[uid]["monthly_details"].setdefault(month_key, {})
                    users[uid]["monthly_details"][month_key].setdefault(
                        r["leave_type"], []
                    ).append({
                        "days": 1,
                        "start": cur_day.strftime("%Y-%m-%d"),
                        "end": cur_day.strftime("%Y-%m-%d")
                    })

                    # ---- total details ----
                    users[uid]["leave_type_details"].setdefault(
                        r["leave_type"], []
                    ).append({
                        "days": 1,
                        "start": cur_day.strftime("%Y-%m-%d"),
                        "end": cur_day.strftime("%Y-%m-%d")
                    })

                cur_day += timedelta(days=1)

        leave_report = list(users.values())

    # =========================================================
    # MONTHLY LEAVE MATRIX (INDEPENDENT)
    # =========================================================
    monthly_matrix = []

    if view == "monthly_matrix":

        first_day = datetime.strptime(
            f"{matrix_year}-{selected_month}-01", "%Y-%m-%d"
        )

        if selected_month == "12":
            last_day = datetime.strptime(
                f"{int(matrix_year)+1}-01-01", "%Y-%m-%d"
            ) - timedelta(days=1)
        else:
            last_day = datetime.strptime(
                f"{matrix_year}-{int(selected_month)+1:02d}-01", "%Y-%m-%d"
            ) - timedelta(days=1)

        users = {}

        # ---- Approved leaves ----
        params = [
            last_day.strftime("%Y-%m-%d"),
            first_day.strftime("%Y-%m-%d")
        ]

        dept_filter = ""
        if selected_dept != "all":
            dept_filter = "AND d.name = ?"
            params.append(selected_dept)

        cur.execute(f"""
            SELECT
                la.user_id,
                u.full_name,
                la.leave_type,
                la.start_date,
                la.end_date
            FROM leave_applications la
            JOIN users u ON u.id = la.user_id
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE la.status = 'Approved'
              AND la.start_date <= ?
              AND la.end_date >= ?
              {dept_filter}
            ORDER BY u.full_name
        """, params)

        rows = cur.fetchall()

        for r in rows:
            uid = r["user_id"]
            users.setdefault(uid, {
                "user_name": r["full_name"],
                "leaves": {}
            })

            start = datetime.strptime(r["start_date"], "%Y-%m-%d")
            end   = datetime.strptime(r["end_date"], "%Y-%m-%d")

            cur_day = max(start, first_day)
            while cur_day <= min(end, last_day):
                users[uid]["leaves"][cur_day.strftime("%d")] = r["leave_type"]
                cur_day += timedelta(days=1)

        # ---- MC records ----
        cur.execute("""
            SELECT
                m.user_id,
                u.full_name,
                m.start_date,
                m.end_date
            FROM mc_records m
            JOIN users u ON u.id = m.user_id
            WHERE m.start_date <= ?
              AND m.end_date >= ?
        """, (
            last_day.strftime("%Y-%m-%d"),
            first_day.strftime("%Y-%m-%d")
        ))

        for m in cur.fetchall():
            uid = m["user_id"]
            users.setdefault(uid, {
                "user_name": m["full_name"],
                "leaves": {}
            })

            start = datetime.strptime(m["start_date"], "%Y-%m-%d")
            end   = datetime.strptime(m["end_date"], "%Y-%m-%d")

            cur_day = max(start, first_day)
            while cur_day <= min(end, last_day):
                users[uid]["leaves"][cur_day.strftime("%d")] = "MC"
                cur_day += timedelta(days=1)

        monthly_matrix = list(users.values())
    # =========================================================
    # RESET UNRELATED FILTERS BY VIEW
    # =========================================================
    view = request.args.get("view", "leave_report")

    if view == "monthly_matrix":
        # reset leave report filters
        selected_department = "all"
        report_year = datetime.now().strftime("%Y")

    elif view == "applied":
        # reset both leave report & matrix filters
        selected_department = "all"
        report_year = datetime.now().strftime("%Y")
        selected_month = datetime.now().strftime("%m")
        selected_dept = "all"
    # =========================================================
    # COMPLETED APPLICATIONS
    # =========================================================
    cur.execute("""
        SELECT
            la.id,
            la.full_name,
            COALESCE(d.name, '') AS department_name,
            la.leave_type,
            la.start_date,
            la.end_date,
            la.status,
            la.approver_name
        FROM leave_applications la
        LEFT JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status IN ('Approved', 'Rejected')
        ORDER BY la.start_date DESC
    """)
    completed = cur.fetchall()

    conn.close()

    return render_template(
        "manage_leaves.html",
        leave_report=leave_report,
        selected_year=report_year,
        selected_department=selected_department,
        monthly_matrix=monthly_matrix,
        selected_month=selected_month,
        selected_year_matrix=matrix_year,
        selected_dept=selected_dept,
        completed=completed,
        departments=departments
    )

MONTHS = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]

def get_leave_matrix_report(year, department_id=None, user_id=None):
    conn = get_db()
    c = conn.cursor()

    sql = """
        SELECT
            u.id AS user_id,
            u.full_name,
            la.leave_type,
            CAST(strftime('%m', la.start_date) AS INTEGER) AS month_no,
            SUM(la.total_days) AS days_used,
            u.entitlement
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        WHERE la.status = 'Approved'
        AND strftime('%Y', la.start_date) = ?
    """
    params = [year]

    if department_id:
        sql += " AND u.department_id = ?"
        params.append(department_id)

    if user_id:
        sql += " AND u.id = ?"
        params.append(user_id)

    sql += """
        GROUP BY u.id, la.leave_type, month_no
        ORDER BY u.full_name, la.leave_type
    """

    rows = c.execute(sql, params).fetchall()
    conn.close()

    return rows

def build_leave_matrix(rows):
    report = {}

    for r in rows:
        uid = r["user_id"]

        if uid not in report:
            report[uid] = {
                "user_id": uid,
                "name": r["full_name"],
                "leaves": []
            }

        leave = next(
            (l for l in report[uid]["leaves"] if l["leave_type"] == r["leave_type"]),
            None
        )

        if not leave:
            leave = {
                "leave_type": r["leave_type"],
                "months": {m: 0.0 for m in MONTHS},
                "total_used": 0.0,
                "remaining": r["entitlement"] or 0
            }
            report[uid]["leaves"].append(leave)

        month_name = MONTHS[r["month_no"] - 1]
        leave["months"][month_name] += r["days_used"]
        leave["total_used"] += r["days_used"]
        leave["remaining"] = max(0, leave["remaining"] - r["days_used"])

    return list(report.values())

@app.route("/download/leave-report/excel")
def download_leave_report_excel():
    import pandas as pd
    from io import BytesIO
    from flask import send_file

    year = request.args.get("year")
    department = request.args.get("department", "all")

    conn = get_db()
    c = conn.cursor()

    query = """
      SELECT 
        u.full_name,
        d.name AS department,
        COALESCE(
          SUM(
            JULIANDAY(l.end_date) - JULIANDAY(l.start_date) + 1
          ), 0
        ) AS used
      FROM users u
      LEFT JOIN departments d ON u.department_id = d.id
      LEFT JOIN leaves l
        ON l.user_id = u.id
       AND l.status = 'Approved'
       AND strftime('%Y', l.start_date) = ?
    """

    params = [year]

    if department != "all":
        query += " WHERE d.name = ?"
        params.append(department)

    query += """
      GROUP BY u.id
      ORDER BY u.full_name
    """

    rows = c.execute(query, params).fetchall()

    df = pd.DataFrame(rows, columns=[
        "Employee Name",
        "Department",
        "Total Leave Used (Days)"
    ])

    output = BytesIO()
    df.to_excel(output, index=False, sheet_name="Leave Report")
    output.seek(0)

    filename = f"LEAVE_REPORT_{department.upper()}_{year}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename
    )

@app.route("/download/leave-report/pdf")
def download_leave_report_pdf():
    from datetime import date

    year = request.args.get("year")
    department = request.args.get("department", "all")

    conn = get_db()
    c = conn.cursor()

    query = """
      SELECT 
        u.full_name,
        d.name AS department,
        COALESCE(
          SUM(
            JULIANDAY(l.end_date) - JULIANDAY(l.start_date) + 1
          ), 0
        ) AS used
      FROM users u
      LEFT JOIN departments d ON u.department_id = d.id
      LEFT JOIN leaves l
        ON l.user_id = u.id
       AND l.status = 'Approved'
       AND strftime('%Y', l.start_date) = ?
    """

    params = [year]

    if department != "all":
        query += " WHERE d.name = ?"
        params.append(department)

    query += """
      GROUP BY u.id
      ORDER BY u.full_name
    """

    rows = c.execute(query, params).fetchall()

    return render_template(
        "leave_report_department_pdf.html",
        year=year,
        department=department,
        printed_date=date.today().strftime("%d-%m-%Y"),
        rows=rows
    )


@app.route("/admin/leave-report/matrix")
@admin_required
def view_leave_matrix():
    year = request.args.get("year", datetime.now().year)
    department = request.args.get("department_id")

    rows = get_leave_matrix_report(year, department)
    matrix = build_leave_matrix(rows)

    return render_template(
        "reports/leave_matrix_preview.html",
        year=year,
        matrix=matrix,
        months=MONTHS
    )

@app.route("/admin/leave-report/matrix/download/excel")
@admin_required
def download_leave_matrix_excel():
    year = request.args.get("year", datetime.now().year)

    rows = get_leave_matrix_report(year)
    data = build_leave_matrix(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    headers = ["ID","Name","Leave Type"] + MONTHS + ["Total Used","Remaining"]
    ws.append(headers)

    for r in data:
        ws.append([
            r["user_id"],
            r["name"],
            r["leave_type"],
            *[r["months"][m] for m in MONTHS],
            r["total_used"],
            r["remaining"]
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=f"Leave_Report_{year}.xlsx"
    )

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

@app.route("/admin/leave-report/matrix/download/pdf")
@admin_required
def download_leave_matrix_pdf():
    year = request.args.get("year", datetime.now().year)

    rows = get_leave_matrix_report(year)
    data = build_leave_matrix(rows)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer)

    table_data = [["ID","Name","Leave Type"] + MONTHS + ["Total","Remaining"]]

    for r in data:
        table_data.append(
            [r["user_id"], r["name"], r["leave_type"]]
            + [r["months"][m] for m in MONTHS]
            + [r["total_used"], r["remaining"]]
        )

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (3,1), (-1,-1), "CENTER"),
    ]))

    doc.build([table])
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Leave_Report_{year}.pdf",
        mimetype="application/pdf"
    )

def get_leave_report(report_type="monthly", month=None):
    conn = get_db()
    c = conn.cursor()

    if month:
        where_clause = "strftime('%Y-%m', la.start_date) = ?"
        params = [month]
    elif report_type == "yearly":
        where_clause = "strftime('%Y', la.start_date) = strftime('%Y','now')"
        params = []
    else:
        where_clause = "strftime('%Y-%m', la.start_date) = strftime('%Y-%m','now')"
        params = []

    c.execute(f"""
        SELECT
            la.user_id,
            la.full_name,
            COALESCE(d.name, '-') AS department_name,
            COUNT(la.id) AS total_application,
            COALESCE(SUM(la.total_days), 0) AS total_days
        FROM leave_applications la
        LEFT JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE {where_clause}
        GROUP BY la.user_id
        ORDER BY la.full_name
    """, params)

    rows = c.fetchall()
    conn.close()
    return rows

def get_all_employee_leaves(view="monthly"):
    conn = get_db()
    c = conn.cursor()

    if view == "weekly":
        where_clause = "date(start_date) >= date('now','-7 day')"
    else:  # monthly
        where_clause = "strftime('%Y-%m', start_date) = strftime('%Y-%m','now')"

    c.execute(f"""
        SELECT
            la.id,
            la.full_name,
            COALESCE(d.name, '-') AS department_name,
            la.leave_type,
            la.start_date,
            la.end_date,
            la.status
        FROM leave_applications la
        LEFT JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE {where_clause}
        ORDER BY la.start_date DESC
    """)

    rows = c.fetchall()
    conn.close()
    return rows


@app.route("/deputygm/dashboard")
@login_required
def deputygm_dashboard():
    # Normalize position
    pos = (session.get("position") or "").strip().upper()

    # Accept ANY Deputy GM variant (HR, ESG, none, etc)
    if not pos.startswith("DEPUTY GENERAL MANAGER"):
        flash("Not allowed.", "danger")
        return redirect(url_for("user_dashboard"))

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT l.*, u.full_name, u.position AS user_position,
               d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.status = 'Pending Check'
        ORDER BY l.id DESC
    """)
    leaves = cur.fetchall()
    conn.close()

    return render_template("dgmdash.html", leaves=leaves)

@app.route("/gm/dashboard")
@login_required
def gm_dashboard():

    pos = (session.get("position") or "").strip().upper()

    # Accept ANY GM variant
    if not pos.startswith("GENERAL MANAGER"):
        flash("Not allowed.", "danger")
        return redirect(url_for("user_dashboard"))

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT l.*, u.full_name, u.position AS user_position,
               d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.status = 'Pending Approval'
        ORDER BY l.id DESC
    """)
    leaves = cur.fetchall()
    conn.close()

    return render_template("gmdash.html", leaves=leaves)

# ===========================
# DGM CHECK ACTION
# ===========================
@app.route("/leave/check/<int:leave_id>", methods=["POST"])
@login_required
def check_leave_action(leave_id):
    pos = (session.get("position") or "").upper()
    if not pos.startswith("DEPUTY GENERAL MANAGER"):
        flash("Not authorized.", "danger")
        return redirect(url_for("user_dashboard"))

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM leave_applications WHERE id=?", (leave_id,))
    leave = cur.fetchone()

    if not leave:
        flash("Leave not found.", "danger")
        conn.close()
        return redirect(url_for("user_dashboard"))

    if leave["status"] != "Pending Check":
        flash("Leave has already been processed.", "warning")
        conn.close()
        return redirect(url_for("user_dashboard"))

    cur.execute("""
        UPDATE leave_applications
        SET status = 'Pending Approval',
            checked_at = ?
        WHERE id = ?
    """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), leave_id))

    conn.commit()
    conn.close()

    flash("Leave checked → sent to General Manager.", "success")
    return redirect(url_for("user_dashboard"))

# ===========================
# GM APPROVE ACTION
# ===========================
@app.route("/leave/approve/<int:leave_id>", methods=["POST"])
@login_required
def approve_leave_action(leave_id):
    pos = (session.get("position") or "").upper()
    if not pos.startswith("GENERAL MANAGER"):
        flash("Not authorized.", "danger")
        return redirect(url_for("user_dashboard"))

    conn = get_db()
    cur = conn.cursor()

    # Fetch leave
    cur.execute("SELECT * FROM leave_applications WHERE id=?", (leave_id,))
    leave = cur.fetchone()

    if not leave:
        flash("Leave not found.", "danger")
        conn.close()
        return redirect(url_for("user_dashboard"))

    # 🧮 Get total leave days
    total_days = int(leave["total_days"] or 0)

    # 🔽 Deduct entitlement from user AFTER approval
    user_id = leave["user_id"]
    cur.execute("SELECT entitlement FROM users WHERE id=?", (user_id,))
    user = cur.fetchone()

    if user:
        remaining = int(user["entitlement"] or 0) - total_days
        if remaining < 0:
            remaining = 0  # avoid negative values

        # Update entitlement
        cur.execute("UPDATE users SET entitlement=? WHERE id=?", (remaining, user_id))

    # 🔓 Approve leave request
    cur.execute("""
        UPDATE leave_applications
        SET status = 'Approved',
            approved_at = ?
        WHERE id = ?
    """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), leave_id))

    conn.commit()
    conn.close()

    flash("Leave approved and deducted successfully.", "success")
    return redirect(url_for("user_dashboard"))

# ===========================
# REJECT ACTION (DGM or GM)
# ===========================
@app.route("/leave/reject/<int:leave_id>", methods=["POST"])
@login_required
def reject_leave_action(leave_id):
    pos = (session.get("position") or "").upper()

    if not (pos.startswith("DEPUTY GENERAL MANAGER") or pos.startswith("GENERAL MANAGER")):
        flash("Not authorized.", "danger")
        return redirect(url_for("user_dashboard"))

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM leave_applications WHERE id=?", (leave_id,))
    leave = cur.fetchone()

    if not leave:
        flash("Leave not found.", "danger")
        conn.close()
        return redirect(url_for("user_dashboard"))

    cur.execute("""
        UPDATE leave_applications
        SET status='Rejected',
            approved_at=?
        WHERE id=?
    """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), leave_id))

    conn.commit()
    conn.close()

    flash("Leave Rejected.", "info")
    return redirect(url_for("user_dashboard"))

@app.route("/admin/leaves/approve/<int:leave_id>", methods=["POST"])
@admin_required
def assign_approver(leave_id):
    """Admin chooses who will approve a leave after it’s checked."""
    approver_position = request.form.get("approver_position")
    approver_department = request.form.get("approver_department")

    if not approver_position:
        flash("Please select approver position.", "warning")
        return redirect(url_for("manage_leaves"))

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        UPDATE leaves
        SET next_approver=?, next_approver_position=?, next_approver_department=?
        WHERE id=?
    """, (approver_position, approver_position, approver_department, leave_id))
    conn.commit()
    conn.close()

    flash(f"Leave assigned to be approved by {approver_position} ({approver_department or 'All Departments'}).", "info")
    return redirect(url_for("manage_leaves"))


@app.route("/leave/action/<int:leave_id>/<action>", methods=["POST"])
@login_required
def update_leave_status(leave_id, action):
    user_id = session["user_id"]

    conn = get_db()
    c = conn.cursor()

    # Fetch leave
    c.execute("SELECT * FROM leaves WHERE id=?", (leave_id,))
    leave = c.fetchone()

    if not leave:
        conn.close()
        flash("Leave not found.", "danger")
        return redirect(url_for("user_dashboard"))

    # --- ONLY ASSIGNED APPROVER CAN APPROVE ---
    if action == "approve":
        if leave["approved_by_user_id"] != user_id:
            flash("You are NOT authorized to approve this leave.", "danger")
            conn.close()
            return redirect(url_for("user_dashboard"))

        # Approve leave
        c.execute("UPDATE leaves SET status='Approved' WHERE id=?", (leave_id,))
        conn.commit()

        flash("Leave approved successfully.", "success")
        conn.close()
        return redirect(url_for("user_dashboard"))

    # --- ONLY CHECKER CAN CHECK OR REJECT ---
    if action == "check":
        if leave["checked_by_user_id"] != user_id:
            flash("You are NOT authorized to check this leave.", "danger")
            conn.close()
            return redirect(url_for("user_dashboard"))

        c.execute("UPDATE leaves SET checked_status='Checked' WHERE id=?", (leave_id,))
        conn.commit()

        flash("Leave checked successfully.", "success")
        conn.close()
        return redirect(url_for("user_dashboard"))

    if action == "reject":
        # Checker OR Approver can reject (depending on position)
        if leave["checked_by_user_id"] != user_id and leave["approved_by_user_id"] != user_id:
            flash("You are NOT authorized to reject this leave.", "danger")
            conn.close()
            return redirect(url_for("user_dashboard"))

        c.execute("UPDATE leaves SET status='Rejected' WHERE id=?", (leave_id,))
        conn.commit()

        flash("Leave rejected.", "info")
        conn.close()
        return redirect(url_for("user_dashboard"))


@app.route('/update-leave-status', methods=['POST'])
@login_required
def update_leave_status_modal():
    """AJAX endpoint for approving/rejecting leave directly from modal."""
    data = request.get_json()
    leave_id = data.get('id')
    new_status = data.get('status')

    if not leave_id or new_status not in ['Approved', 'Rejected', 'Pending']:
        return jsonify({'success': False, 'error': 'Invalid request'}), 400

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM leaves WHERE id=?", (leave_id,))
    leave = c.fetchone()

    if not leave:
        conn.close()
        return jsonify({'success': False, 'error': 'Leave not found'}), 404

    c.execute("UPDATE leaves SET status=? WHERE id=?", (new_status, leave_id))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'status': new_status})


@app.route("/admin/holidays", methods=["GET", "POST"])
@admin_required
def holidays():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        day = request.form.get("date", "").strip()
        if name and day:
            c.execute("INSERT INTO holidays (name, date) VALUES (?,?)", (name, day))
            conn.commit()
            flash("Holiday added.", "success")
        else:
            flash("Please provide name and date.", "danger")
    c.execute("SELECT * FROM holidays ORDER BY date")
    rows = c.fetchall()
    conn.close()
    return render_template("holidays.html", holidays=rows)

@app.route("/admin/departments", methods=["GET", "POST"])
@admin_required
def manage_departments():
    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        dept_name = request.form.get("name", "").strip()
        if dept_name:
            try:
                c.execute("INSERT INTO departments (name) VALUES (?)", (dept_name,))
                conn.commit()
                flash("Department added successfully.", "success")
            except sqlite3.IntegrityError:
                flash("Department already exists.", "warning")
            finally:
                conn.close()

            # 🟢 FIX: Redirect back to Manage Users instead of Departments page
            return redirect(url_for("manage_users"))

    # GET method (only if user visits /admin/departments directly)
    c.execute("SELECT * FROM departments ORDER BY name")
    departments = c.fetchall()
    conn.close()

    return render_template("manage_departments.html", departments=departments)


@app.route("/admin/departments/delete/<int:dept_id>", methods=["POST"])
@admin_required
def delete_department(dept_id):
    """Delete department and stay on Manage Users dashboard."""
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM departments WHERE id=?", (dept_id,))
    conn.commit()
    conn.close()
    flash("Department deleted successfully.", "info")

    # ✅ Stay on Manage Users dashboard
    return redirect(url_for("manage_users"))


# ---------------------- User Views ----------------------
@app.route("/user/dashboard")
@login_required
def user_dashboard():
    user_id = session["user_id"]

    conn = get_db()
    c = conn.cursor()

    # 🟦 Get user position + normalize
    c.execute("SELECT position FROM users WHERE id=?", (user_id,))
    my_pos = (c.fetchone()["position"] or "").upper().strip()

    # 🟨 TO CHECK (For Deputy General Manager ONLY)
    c.execute("""
        SELECT l.*, u.full_name, u.position, d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.status = 'Pending Check'
        ORDER BY l.id DESC
    """)
    to_check = c.fetchall()

    # 🟥 TO APPROVE (For General Manager ONLY)
    c.execute("""
        SELECT l.*, u.full_name, u.position, d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.status = 'Pending Approval'
        ORDER BY l.id DESC
    """)
    to_approve = c.fetchall()

    # 🟩 MY OWN LEAVES
    c.execute("""
        SELECT l.*, u.full_name, d.name AS department_name
        FROM leave_applications l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.user_id=?
        ORDER BY l.id DESC
    """, (user_id,))
    my_leaves = c.fetchall()
    
    # ===================== MY MC RECORDS =====================
    c.execute("""
        SELECT mc_number, start_date, end_date, pdf_path, created_at
        FROM mc_records
        WHERE user_id = ?
        ORDER BY created_at DESC
    """, (user_id,))
    my_mc = c.fetchall()


    # 🚫 Hide notifications from normal users
    if not my_pos.startswith("DEPUTY GENERAL MANAGER"):
        to_check = []
    if not my_pos.startswith("GENERAL MANAGER"):
        to_approve = []

    conn.close()

    return render_template(
    "user_dashboard.html",
    my_leaves=my_leaves,
    to_check=to_check,
    to_approve=to_approve,
    total_pending=len(to_check) + len(to_approve),
    my_pos=my_pos,
    my_mc=my_mc )


# ✅ Auto workflow mapping based on leave approval rules
AUTO_ASSIGN = {
    "ASSISTANT": {"checker": "DEPUTY GENERAL MANAGER", "approver": "GENERAL MANAGER"},
    "EXECUTIVE": {"checker": "DEPUTY GENERAL MANAGER", "approver": "GENERAL MANAGER"},
    "HR ASSISTANT": {"checker": "DEPUTY GENERAL MANAGER", "approver": "GENERAL MANAGER"},
    "HR EXECUTIVE": {"checker": "DEPUTY GENERAL MANAGER", "approver": "GENERAL MANAGER"},
    "HR MANAGER": {"checker": "DEPUTY GENERAL MANAGER", "approver": "GENERAL MANAGER"},
    "FELO SMJ": {"checker": "DEPUTY GENERAL MANAGER", "approver": "GENERAL MANAGER"},

    "DEPUTY GENERAL MANAGER": {"checker": None, "approver": "GENERAL MANAGER"},
    "GENERAL MANAGER": {"checker": None, "approver": "CEO"},
    "CEO": {"checker": None, "approver": None}
}


# Fallback workflow for unknown positions (FELO, INTERN, etc)
FALLBACK_CHECKER = "DEPUTY GENERAL MANAGER - HR", "Deputy General Manager"
FALLBACK_APPROVER = "GENERAL MANAGER - HR & ESG"

@app.route("/user/upload_mc", methods=["POST"])
@login_required
def user_upload_mc():
    mc_number = request.form.get("mc_number", "").strip()
    mc_start  = request.form.get("mc_start") or None
    mc_end    = request.form.get("mc_end") or None
    file      = request.files.get("mc_file")

    if not file or file.filename == "":
        flash("MC file required.", "warning")
        return redirect(url_for("user_dashboard"))

    filename = secure_filename(file.filename)
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in {"pdf", "png", "jpg", "jpeg"}:
        flash("Invalid MC file type.", "danger")
        return redirect(url_for("user_dashboard"))

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    save_name = f"mc_{session['user_id']}_{ts}_{filename}"
    file.save(os.path.join(LEAVE_UPLOAD_FOLDER, save_name))

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO mc_records
        (user_id, mc_number, start_date, end_date, pdf_path, uploaded_by, created_at)
        VALUES (?,?,?,?,?,?,?)
    """, (
        session["user_id"],
        mc_number,
        mc_start,
        mc_end,
        save_name,
        session["user_id"],
        datetime.utcnow().isoformat()
    ))

    c.execute(
        "UPDATE users SET availability='MC' WHERE id=?",
        (session["user_id"],)
    )

    conn.commit()
    conn.close()

    flash("Medical Certificate uploaded successfully.", "success")
    return redirect(url_for("user_dashboard"))


def normalize_position(pos):
    if not pos:
        return ""
    # remove extra spaces, unify dashes and upper-case
    return " ".join(pos.replace("–", "-").replace("—", "-").strip().upper().split())



@app.route("/apply_leave", methods=["GET", "POST"])
@login_required
def apply_leave():
    user_id = session.get("user_id")
    full_name = session.get("full_name")
    raw_position = (session.get("position") or "").strip().upper()

    # Normalize positions
    if "DEPUTY GENERAL MANAGER" in raw_position:
        pos_upper = "DEPUTY GENERAL MANAGER"
    elif "GENERAL MANAGER" in raw_position:
        pos_upper = "GENERAL MANAGER"
    else:
        pos_upper = raw_position

    # ================= POST =================
    if request.method == "POST":

        leave_type = request.form.get("leave_type")
        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        reason = request.form.get("reason")

        # Calculate working days
        try:
            total_days = calculate_working_days(start_date, end_date)
        except:
            total_days = 1

        # ================= WORKFLOW RULES =================

        # Default checker & approver from form
        checker_name = request.form.get("checker")
        approver_name = request.form.get("approver")

        # Staff → checker = DGM, approver = GM
        if pos_upper not in ["DEPUTY GENERAL MANAGER", "GENERAL MANAGER"]:
            checker_name = "Deputy General Manager"
            approver_name = "General Manager"
            status = "Pending Check"

        # DGM → no checker → approver = GM
        elif pos_upper == "DEPUTY GENERAL MANAGER":
            checker_name = None
            approver_name = "General Manager"
            status = "Pending Approval"

        # GM → no checker → approver = CEO
        elif pos_upper == "GENERAL MANAGER":
            checker_name = None
            approver_name = "CEO"
            status = "Pending Approval"

        # CEO auto approve
        elif pos_upper == "CEO":
            checker_name = None
            approver_name = None
            status = "Approved"

        # Save supporting doc
        support_doc = None
        file = request.files.get("support_doc")
        if file and file.filename.strip():
            if allowed_leave_file(file.filename):
                original = secure_filename(file.filename)
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                support_doc = f"{user_id}_{timestamp}_{original}"
                file.save(os.path.join(LEAVE_UPLOAD_FOLDER, support_doc))

        # ================= INSERT =================
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO leave_applications
            (user_id, full_name, position, leave_type, start_date, end_date, total_days,
             reason, status, checker_name, approver_name, support_doc, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user_id, full_name, pos_upper, leave_type, start_date, end_date, total_days,
            reason, status, checker_name, approver_name, support_doc,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))

        conn.commit()
        conn.close()

        flash("🎉 Leave submitted successfully!", "success")
        return redirect(url_for("user_dashboard"))

    # ================= GET =================
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT entitlement FROM users WHERE id=?", (user_id,))
    row = cur.fetchone()
    conn.close()

    return render_template(
        "apply_leave.html",
        full_name=full_name,
        current_date=datetime.now().strftime("%d/%m/%Y"),
        position=session.get("position"),
        remaining_leave=row["entitlement"] if row else 0
    )


@app.route("/leave/<int:leave_id>/approve", methods=["POST"])
@login_required
def approve_leave(leave_id):
    if session.get("position") not in ("GENERAL MANAGER - HR & ESG", "GENERAL MANAGER"):
        flash("You are not authorized to perform this action.", "danger")
        return redirect(url_for("dashboard"))

    conn = get_db()
    cur = conn.cursor()

    approved_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cur.execute(
        """
        UPDATE leave_applications
        SET status = 'Approved',
            approved_at = ?
        WHERE id = ?
        """,
        (approved_at, leave_id),
    )

    conn.commit()
    conn.close()

    flash("Leave has been approved.", "success")
    return redirect(url_for("gm_dashboard"))


@app.route("/leave/<int:leave_id>/reject", methods=["POST"])
@login_required
def reject_leave(leave_id):
    if session.get("position") not in ("GENERAL MANAGER - HR & ESG", "GENERAL MANAGER"):
        flash("You are not authorized to perform this action.", "danger")
        return redirect(url_for("dashboard"))

    conn = get_db()
    cur = conn.cursor()

    approved_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cur.execute(
        """
        UPDATE leave_applications
        SET status = 'Rejected',
            approved_at = ?
        WHERE id = ?
        """,
        (approved_at, leave_id),
    )

    conn.commit()
    conn.close()

    flash("Leave has been rejected.", "info")
    return redirect(url_for("gm_dashboard"))


@app.route("/leave/<int:leave_id>/check", methods=["POST"])
@login_required
def check_leave(leave_id):
    # Hanya Deputy GM boleh check
    if session.get("position") != "DEPUTY GENERAL MANAGER - HR":
        flash("You are not authorized to perform this action.", "danger")
        return redirect(url_for("dashboard"))

    conn = get_db()
    cur = conn.cursor()

    checked_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Bila sudah check → tukar status ke "Pending Approval" untuk GM
    cur.execute(
        """
        UPDATE leave_applications
        SET status = 'Pending Approval',
            checked_at = ?
        WHERE id = ?
        """,
        (checked_at, leave_id),
    )

    conn.commit()
    conn.close()

    flash("Leave has been checked and sent to General Manager for approval.", "success")
    return redirect(url_for("deputygm_dashboard"))

@app.route("/debug/positions")
def debug_positions():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, full_name, position FROM users")
    rows = c.fetchall()
    conn.close()

    output = "<h3>Users & Positions in DB</h3><ul>"
    for r in rows:
        output += f"<li>ID: {r['id']}, Name: {r['full_name']}, Position Stored: '{r['position']}'</li>"
    output += "</ul>"
    return output

@app.route("/calendar")
@login_required
def calendar():
    conn = get_db()
    c = conn.cursor()

    # Fetch holidays (still global)
    c.execute("SELECT name, date FROM holidays ORDER BY date")
    holidays = c.fetchall()

    # 🔹 Upcoming leaves for this user only
    c.execute("""
        SELECT l.id, l.start_date, l.end_date, u.full_name, l.leave_type, l.status
        FROM leaves l
        JOIN users u ON u.id = l.user_id
        WHERE l.user_id = ?
        AND date(l.end_date) >= date('now')
        ORDER BY l.start_date ASC
    """, (session["user_id"],))
    leaves = c.fetchall()

    # 🔹 Leave history (past leaves for this user only)
    c.execute("""
        SELECT l.id, l.start_date, l.end_date, u.full_name, l.leave_type, l.status
        FROM leaves l
        JOIN users u ON u.id = l.user_id
        WHERE l.user_id = ?
        AND date(l.end_date) < date('now')
        ORDER BY l.start_date DESC
    """, (session["user_id"],))
    leave_history = c.fetchall()

    conn.close()
    return render_template(
        "calendar.html",
        holidays=holidays,
        leaves=leaves,
        leave_history=leave_history
    )

@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        password = request.form.get("password", "").strip()

        # ▶ Update Full Name
        if full_name:
            c.execute("UPDATE users SET full_name=? WHERE id=?", 
                      (full_name, session["user_id"]))
            session["full_name"] = full_name  # Update session UI immediately

        # ▶ Update Password ONLY if user typed something
        if password:
            c.execute("UPDATE users SET password_hash=? WHERE id=?", 
                      (generate_password_hash(password), session["user_id"]))

        conn.commit()
        conn.close()
        flash("Profile updated successfully.", "success")

        # ⛔ DO NOT render directly — REDIRECT to avoid undefined 'user'
        return redirect(url_for("profile"))

    # ================================
    # GET: Fetch user + department info
    # ================================
    c.execute("""
        SELECT 
            u.*, d.name AS department_name
        FROM users u
        LEFT JOIN departments d ON d.id = u.department_id
        WHERE u.id=?
    """, (session["user_id"],))
    
    user = c.fetchone()
    conn.close()

    return render_template("profile.html", user=user)


@app.route("/update_profile_photo", methods=["POST"])
@login_required
def update_profile_photo():
    file = request.files.get("profile_photo")

    if not file or file.filename == "":
        flash("No file selected.", "warning")
        return redirect(url_for("profile"))

    if not allowed_photo(file.filename):
        flash("Invalid file type. Please upload JPG, PNG or GIF.", "danger")
        return redirect(url_for("profile"))

    ext = file.filename.rsplit(".", 1)[1].lower()
    filename = secure_filename(f"user_{session['user_id']}.{ext}")
    save_path = os.path.join(app.config["PROFILE_UPLOAD_FOLDER"], filename)
    file.save(save_path)

    # update DB
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE users SET profile_photo=? WHERE id=?", (filename, session["user_id"]))
    conn.commit()
    conn.close()

    # ⭐ update session so navbar shows image
    session['profile_photo'] = filename

    flash("Profile photo updated successfully.", "success")
    return redirect(url_for("profile"))

@app.route('/delete_profile_photo', methods=['POST'])
@login_required
def delete_profile_photo():
    """Delete user's current profile photo and restore default avatar."""
    user_id = session['user_id']

    conn = get_db()
    c = conn.cursor()

    # Get current photo filename
    c.execute("SELECT profile_photo FROM users WHERE id=?", (user_id,))
    result = c.fetchone()

    if result and result["profile_photo"]:
        photo_path = os.path.join(app.config["UPLOAD_FOLDER"], result["profile_photo"])
        if os.path.exists(photo_path):
            os.remove(photo_path)  # delete the image file

    # Set to NULL (system will show avatar)
    c.execute("UPDATE users SET profile_photo=NULL WHERE id=?", (user_id,))
    conn.commit()
    conn.close()

    flash("Profile photo deleted successfully. Default avatar restored.", "success")
    return redirect(url_for("profile"))


@app.route("/settings", methods=["GET", "POST"])
@admin_required
def settings():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        org_name = request.form.get("org_name", "J-Leave App")
        theme = request.form.get("theme", "blue")
        c.execute("INSERT INTO settings (key, value) VALUES ('org_name', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (org_name,))
        c.execute("INSERT INTO settings (key, value) VALUES ('theme', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (theme,))
        conn.commit()
        conn.close()
        flash("Settings saved.", "success")
        return redirect(url_for("settings"))
    c.execute("SELECT key, value FROM settings")
    settings_rows = {row["key"]: row["value"] for row in c.fetchall()}
    conn.close()
    return render_template("settings.html", settings=settings_rows)
@app.route("/admin/leaves/assign-checker/<int:leave_id>", methods=["POST"])
@admin_required
def assign_checker(leave_id):
    """Admin assigns a specific user to check this leave."""
    checker_id = request.form.get("checker_id")
    if not checker_id:
        flash("Please select a checker.", "warning")
        return redirect(url_for("manage_leaves"))

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT full_name, position FROM users WHERE id=?", (checker_id,))
    checker = c.fetchone()

    if not checker:
        flash("Invalid checker selected.", "danger")
        conn.close()
        return redirect(url_for("manage_leaves"))

    c.execute("""
        UPDATE leaves 
        SET checked_by_user_id=?, checked_by_position=?, checked_status='Pending'
        WHERE id=?
    """, (checker_id, checker["position"], leave_id))
    conn.commit()
    conn.close()

    flash(f"Checker assigned: {checker['full_name']} ({checker['position']}).", "success")
    return redirect(url_for("manage_leaves"))


@app.route("/admin/leaves/assign-approver/<int:leave_id>", methods=["POST"])
@admin_required
def assign_approver_user(leave_id):
    """Admin assigns a specific user to approve this leave."""
    approver_id = request.form.get("approver_id")
    if not approver_id:
        flash("Please select an approver.", "warning")
        return redirect(url_for("manage_leaves"))

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT full_name, position FROM users WHERE id=?", (approver_id,))
    approver = c.fetchone()

    if not approver:
        flash("Invalid approver selected.", "danger")
        conn.close()
        return redirect(url_for("manage_leaves"))

    c.execute("""
        UPDATE leaves 
        SET approved_by_user_id=?, next_approver=?, next_approver_position=?, status='Pending'
        WHERE id=?
    """, (approver_id, approver["full_name"], approver["position"], leave_id))
    conn.commit()
    conn.close()

    flash(f"Approver assigned: {approver['full_name']} ({approver['position']}).", "success")
    return redirect(url_for("manage_leaves"))

@app.route("/toggle-theme")
def toggle_theme():
    current = session.get("theme_mode", "light")
    session["theme_mode"] = "dark" if current == "light" else "light"
    return redirect(request.referrer or url_for("user_dashboard"))


# --------------- simple API for live clock (optional) ---------------
# @app.route("/api/server-time")
# def server_time():
#     return jsonify({"server_time": datetime.now().isoformat()})
@app.route("/api/leave-trend")
def leave_trend_api():
    view = request.args.get("view", "weekly")
    if view == "monthly":
        labels = ["Jan", "Feb", "Mar", "Apr"]
        annual = [5, 8, 6, 7]
        sick = [2, 3, 1, 4]
    else:
        labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        annual = [2, 1, 3, 1, 0, 0, 2]
        sick = [0, 1, 0, 1, 1, 0, 0]
    return {"labels": labels, "annual": annual, "sick": sick}

@app.route("/leave/details/<int:leave_id>")
@login_required
def leave_details(leave_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT 
            l.*, 
            u.full_name, u.position, u.email, u.phone, u.address,
            d.name AS department_name
        FROM leave_applications l
        JOIN users u ON l.user_id = u.id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.id=?
    """, (leave_id,))
    leave = c.fetchone()
    conn.close()

    if not leave:
        flash("Leave record not found.", "danger")
        return redirect(url_for("user_dashboard"))

    return render_template("leave_details.html", leave=leave, datetime=datetime.now())


@app.route("/api/leave/<int:leave_id>")
@login_required
def api_leave_details(leave_id):
    """Return leave details as JSON for modal display."""
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT 
            l.*, 
            u.full_name, u.position, u.email, u.phone, u.address,
            d.name AS department_name
        FROM leaves l
        JOIN users u ON l.user_id = u.id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.id=?
    """, (leave_id,))
    leave = c.fetchone()
    conn.close()

    if not leave:
        return jsonify({"error": "Leave not found"}), 404

    return jsonify(dict(leave))

@app.route("/debug/leaves")
def debug_leaves():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT id, full_name, position, status, checker_name, approver_name
        FROM leave_applications
        ORDER BY id DESC
    """)
    rows = c.fetchall()
    conn.close()

    out = "<h2>Leave Applications DEBUG</h2><table border=1 cellpadding=5>"
    out += "<tr><th>ID</th><th>Name</th><th>Position</th><th>Status</th><th>Checker</th><th>Approver</th></tr>"
    for r in rows:
        out += f"<tr><td>{r['id']}</td><td>{r['full_name']}</td><td>{r['position']}</td><td>{r['status']}</td><td>{r['checker_name']}</td><td>{r['approver_name']}</td></tr>"
    out += "</table>"
    return out

def get_departments():
    conn = sqlite3.connect("database.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT DISTINCT department
        FROM users
        WHERE department IS NOT NULL AND department != ''
        ORDER BY department
    """)

@app.route("/leave_docs/<path:filename>")
@login_required
def leave_docs(filename):
    # Semua user login boleh view, kalau nak stricter, check role sini
    return send_from_directory(LEAVE_UPLOAD_FOLDER, filename)

import smtplib
from email.mime.text import MIMEText

def send_notification(user_id, message):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT email, phone FROM users WHERE id=?", (user_id,))
    user = c.fetchone()
    conn.close()

    if not user:
        return

    if user["email"]:
        send_email(user["email"], "J-Leave Notification", message)

    if user["phone"]:
        send_whatsapp(user["phone"], message)

def send_email(to_email, subject, message):
    sender = "noreply@jleave.com"
    msg = MIMEText(message)
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = to_email

    try:
        smtp = smtplib.SMTP("smtp.gmail.com", 587)
        smtp.starttls()
        smtp.login("YOUR_EMAIL@gmail.com", "YOUR_APP_PASSWORD")
        smtp.sendmail(sender, [to_email], msg.as_string())
        smtp.quit()
    except Exception as e:
        print("Email error:", e)


def send_whatsapp(phone, message):
    print(f"WhatsApp sent to {phone}: {message}")
    # ✅ Integrate Twilio / Fonnte later here

@app.route("/leave/file/<filename>")
@login_required
def leave_file(filename):
    return send_from_directory(LEAVE_UPLOAD_FOLDER, filename)

@app.route("/leave/<int:leave_id>/download/pdf")
@login_required
def download_leave_pdf(leave_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT l.*, u.full_name, u.position, u.email, u.phone, u.address,
               d.name AS department_name
        FROM leave_applications l
        JOIN users u ON l.user_id = u.id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE l.id=?
    """, (leave_id,))
    leave = c.fetchone()
    conn.close()

    if not leave:
        flash("Leave record not found.", "danger")
        return redirect(url_for("user_dashboard"))

    # Render HTML template
    html = render_template("leave_details.html", leave=leave, datetime=datetime.now())

    # Convert HTML → PDF
    pdf = HTML(string=html, base_url=request.host_url).write_pdf()

    return send_file(
        io.BytesIO(pdf),
        as_attachment=True,
        download_name=f"leave_application_{leave_id}.pdf",
        mimetype="application/pdf"
    )

@app.route("/forgot", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip()

        if not email:
            flash("Please enter your registered email.", "danger")
            return redirect(url_for("forgot_password"))

        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT id FROM users WHERE email=?", (email,))
        user = c.fetchone()

        if not user:
            conn.close()
            flash("Email not found in system.", "danger")
            return redirect(url_for("forgot_password"))

        # Generate token & expiry (1 hour)
        token = str(uuid4())
        expiry = (datetime.utcnow() + timedelta(hours=1)).isoformat()

        c.execute("""
            UPDATE users
            SET reset_token = ?, reset_token_expiry = ?
            WHERE email = ?
        """, (token, expiry, email))
        conn.commit()
        conn.close()

        reset_link = f"{request.host_url}reset/{token}"
        send_email(email, "Reset Your J-Leave Account",
                   f"Click the link below to reset your username/password:\n\n{reset_link}\n\nThis link will expire in 1 hour.")

        # reset login attempts so user can try again after reset
        session["login_attempts"] = 0

        flash("Reset link sent. Please check your email.", "info")
        return redirect(url_for("login"))

    return render_template("forgot_password.html")
    

@app.route("/reset/<token>", methods=["GET", "POST"])
def reset_password(token):
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT id, reset_token_expiry
        FROM users
        WHERE reset_token = ?
    """, (token,))
    user = c.fetchone()

    if not user:
        conn.close()
        flash("Invalid or expired reset link.", "danger")
        return redirect(url_for("login"))

    expiry_str = user["reset_token_expiry"]
    try:
        expiry = datetime.fromisoformat(expiry_str)
    except Exception:
        conn.close()
        flash("Invalid token data. Please request a new reset link.", "danger")
        return redirect(url_for("forgot_password"))

    if datetime.utcnow() > expiry:
        conn.close()
        flash("Reset link has expired. Please request again.", "warning")
        return redirect(url_for("forgot_password"))

    if request.method == "POST":
        new_username = (request.form.get("username") or "").strip()
        p1 = request.form.get("password") or ""
        p2 = request.form.get("password2") or ""

        if len(p1) < 6:
            flash("Password must be at least 6 characters.", "danger")
            return redirect(request.url)

        if p1 != p2:
            flash("Passwords do not match.", "danger")
            return redirect(request.url)

        updates = []
        params = []

        # optional username change
        if new_username:
            updates.append("username = ?")
            params.append(new_username)

        # always update password
        updates.append("password_hash = ?")
        params.append(generate_password_hash(p1))

        # clear token
        updates.append("reset_token = NULL")
        updates.append("reset_token_expiry = NULL")

        set_clause = ", ".join(updates)

        params.append(user["id"])

        c.execute(f"""
            UPDATE users
            SET {set_clause}
            WHERE id = ?
        """, params)

        conn.commit()
        conn.close()

        flash("Details updated successfully. You may now login.", "success")
        return redirect(url_for("login"))

    # GET – render form
    conn.close()
    return render_template("reset_password.html")

from flask import send_file
import io

@app.route("/export/leave/excel")
def export_excel():
    mode = request.args.get("mode", "weekly")
    data = get_leave_data(mode)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    ws.append(["Employee Name", "Leave Type", "Start Date", "End Date", "Status"])

    for row in data:
        ws.append(row)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"LeaveReport_{mode}.xlsx"
    return send_file(file_stream, as_attachment=True, download_name=filename)

#Downloadedable leave report
# INDIVIDUAL LEAVE REPORT

from datetime import datetime, timedelta
from flask import render_template, request

@app.route("/admin/leave-report/employee/<int:user_id>/view")
@admin_required
def view_individual_leave_report(user_id):

    year = request.args.get("year")
    if not year:
        year = str(datetime.now().year)

    conn = get_db()
    c = conn.cursor()

    # ================= EMPLOYEE =================
    c.execute("""
        SELECT u.full_name, u.position, u.entitlement, d.name AS department
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.id = ?
    """, (user_id,))
    emp = c.fetchone()

    # ================= APPROVED LEAVES =================
    c.execute("""
        SELECT leave_type, start_date, end_date, total_days
        FROM leave_applications
        WHERE user_id = ?
          AND status = 'Approved'
          AND strftime('%Y', start_date) = ?
        ORDER BY start_date
    """, (user_id, year))
    approved_leaves = c.fetchall()

    # ================= MC =================
    c.execute("""
        SELECT start_date, end_date
        FROM mc_records
        WHERE user_id = ?
          AND strftime('%Y', start_date) = ?
        ORDER BY start_date
    """, (user_id, year))
    mc_records = c.fetchall()

    conn.close()

    # ================= MONTHLY BY TYPE =================
    MONTHS = ["JAN","FEB","MAR","APR","MAY","JUN",
              "JUL","AUG","SEP","OCT","NOV","DEC"]

    monthly = {m: {} for m in MONTHS}

    for l in approved_leaves:
        if l["leave_type"] == "MC":
            continue

        start = datetime.strptime(l["start_date"], "%Y-%m-%d").date()
        end   = datetime.strptime(l["end_date"], "%Y-%m-%d").date()

        d = start
        while d <= end:
            if d.year == int(year) and d.weekday() < 5:
                m = MONTHS[d.month - 1]
                monthly[m][l["leave_type"]] = monthly[m].get(l["leave_type"], 0) + 1
            d += timedelta(days=1)

    # ================= SUMMARY =================
    entitled = emp["entitlement"] or 0
    used = sum(sum(v.values()) for v in monthly.values())
    balance = max(0, entitled - used)

    summary = {
        "entitled": entitled,
        "used": used,
        "balance": balance
    }

    # ================= LEAVE TYPES =================
    leave_types = sorted(
        {l["leave_type"] for l in approved_leaves if l["leave_type"] != "MC"}
    )

    return render_template(
        "reports/individual_leave_report.html",
        employee={
            "name": emp["full_name"],
            "position": emp["position"]
        },
        department=emp["department"],
        year=year,
        printed_date=datetime.now().strftime("%d %b %Y"),
        summary=summary,
        monthly=monthly,
        leave_types=leave_types,
        approved_leaves=approved_leaves,
        mc_records=mc_records
    )


from weasyprint import HTML
from datetime import datetime

@app.route("/leave-report/individual/<int:user_id>/pdf")
def download_individual_leave_report_pdf(user_id):
    year = request.args.get("year", datetime.now().year)

    data = build_individual_leave_report(user_id, year)

    html = render_template(
        "reports/individual_leave_report.html",
        **data,
        printed_date=datetime.now().strftime("%d %b %Y")
    )

    pdf = HTML(string=html).write_pdf()

    return Response(
        pdf,
        headers={
            "Content-Type": "application/pdf",
            "Content-Disposition": f"attachment; filename=leave_{user_id}_{year}.pdf"
        }
    )

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from flask import send_file
import io

@app.route("/leave-report/individual/<int:user_id>/excel")
def download_individual_leave_report_excel(user_id):
    year = request.args.get("year", datetime.now().year)

    data = build_individual_leave_report(user_id, year)

    return generate_individual_leave_excel(data, f"leave_{user_id}_{year}.xlsx")


def build_individual_leave_report(user_id, year):
    conn = get_db()
    cur = conn.cursor()

    # Employee info
    cur.execute("""
        SELECT 
            u.full_name,
            u.position,
            d.name AS department
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.id = ?
    """, (user_id,))
    emp = cur.fetchone()

    if not emp:
        conn.close()
        raise ValueError("Employee not found")

    monthly = {m: 0 for m in [
        "January","February","March","April","May","June",
        "July","August","September","October","November","December"
    ]}

    cur.execute("""
        SELECT start_date, end_date
        FROM leave_applications
        WHERE user_id = ?
          AND status = 'Approved'
          AND strftime('%Y', start_date) = ?
    """, (user_id, str(year)))

    leaves = cur.fetchall()
    conn.close()

    for l in leaves:
        start = datetime.strptime(l["start_date"], "%Y-%m-%d").date()
        end   = datetime.strptime(l["end_date"], "%Y-%m-%d").date()

        for d in daterange(start, end):
            if d.year == int(year) and d.weekday() < 5:
                monthly[d.strftime("%B")] += 1

    used = sum(monthly.values())
    entitled = 14

    return {
        "employee": {
            "name": emp["full_name"],
            "department": emp["department"],
            "position": emp["position"]
        },
        "monthly": monthly,
        "summary": {
            "entitled": entitled,
            "used": used,
            "balance": max(0, entitled - used)
        }
    }

from collections import defaultdict
from datetime import datetime, timedelta

def daterange(start, end):
    for n in range((end - start).days + 1):
        yield start + timedelta(n)

def build_employee_leave_matrix(leaves, year):
    months = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']

    from collections import defaultdict
    monthly = {m: defaultdict(int) for m in months}
    monthly_details = {m: defaultdict(list) for m in months}

    total_used = 0

    for l in leaves:
        # ✅ FIX 1
        if l["status"] != "Approved":
            continue

        # ✅ FIX 2
        start = datetime.strptime(l["start_date"], "%Y-%m-%d").date()
        end   = datetime.strptime(l["end_date"], "%Y-%m-%d").date()

        for d in daterange(start, end):
            if d.year != int(year):
                continue

            m = months[d.month - 1]

            # ✅ FIX 3
            monthly[m][l["leave_type"]] += 1

            monthly_details[m][l["leave_type"]].append({
                "date": d.strftime("%d-%m-%Y"),
                "days": 1
            })

            # ✅ FIX 4
            if l["leave_type"] != "MC":
                total_used += 1

    return {
        "monthly": monthly,
        "monthly_details": monthly_details,
        "total_used": total_used
    }
    
from flask import render_template, request, send_file
from datetime import datetime
import pandas as pd
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet

@app.route("/team_leave_report")
@admin_required
def team_leave_report():

    department = request.args.get("department", "").strip()

    conn = get_db()
    c = conn.cursor()

    # ================= DEPARTMENT LIST =================
    c.execute("SELECT name FROM departments ORDER BY name")
    departments = [r["name"] for r in c.fetchall()]

    report = []

    if department:
        c.execute("""
            SELECT
                u.id AS user_id,
                u.full_name,
                u.position,
                u.entitlement,
                la.leave_type,
                la.start_date,
                la.end_date,
                la.total_days
            FROM leave_applications la
            JOIN users u ON u.id = la.user_id
            LEFT JOIN departments d ON u.department_id = d.id
            WHERE la.status = 'Approved'
              AND d.name = ?
            ORDER BY u.full_name, la.start_date
        """, (department,))

        rows = c.fetchall()

        employees = {}

        for r in rows:
            uid = r["user_id"]

            if uid not in employees:
                employees[uid] = {
                    "name": r["full_name"],
                    "position": r["position"],
                    "entitlement": r["entitlement"] or 0,
                    "used": 0,
                    "leaves": []
                }

            # MC does NOT deduct entitlement
            if r["leave_type"] != "MC":
                employees[uid]["used"] += r["total_days"] or 0

            employees[uid]["leaves"].append({
                "leave_type": r["leave_type"],
                "date": f"{r['start_date']} → {r['end_date']}",
                "days": r["total_days"] or 0
            })

        # Final formatting
        for emp in employees.values():
            report.append({
                "name": emp["name"],
                "position": emp["position"],
                "leaves": emp["leaves"],
                "used": emp["used"],
                "balance": max(0, emp["entitlement"] - emp["used"])
            })

    conn.close()

    return render_template(
        "team_leave_report.html",
        departments=departments,
        selected_department=department,
        report=report
    )

@app.route("/team_leave_excel")
@admin_required
def team_leave_excel():

    department = request.args.get("department")
    if not department:
        flash("Please select a department.", "warning")
        return redirect(url_for("team_leave_report"))

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT
            u.full_name,
            u.position,
            la.leave_type,
            la.start_date,
            la.end_date,
            la.total_days
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status = 'Approved'
          AND d.name = ?
        ORDER BY u.full_name, la.start_date
    """, (department,))

    rows = c.fetchall()
    conn.close()

    data = []
    for r in rows:
        data.append([
            r["full_name"],
            r["position"],
            r["leave_type"],
            f"{r['start_date']} → {r['end_date']}",
            r["total_days"]
        ])

    import pandas as pd
    output = io.BytesIO()
    df = pd.DataFrame(data, columns=[
        "Employee Name",
        "Position",
        "Leave Type",
        "Date",
        "Days"
    ])
    df.to_excel(output, index=False)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"Team_Leave_Report_{department}.xlsx"
    )


@app.route("/team_leave_pdf")
@admin_required
def team_leave_pdf():

    department = request.args.get("department")
    if not department:
        flash("Please select a department.", "warning")
        return redirect(url_for("team_leave_report"))

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT
            u.full_name,
            u.position,
            la.leave_type,
            la.start_date,
            la.end_date,
            la.total_days
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status = 'Approved'
          AND d.name = ?
        ORDER BY u.full_name, la.start_date
    """, (department,))

    rows = c.fetchall()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(f"<b>Team Leave Report – {department}</b>", styles["Title"])
    ]

    table_data = [["Name", "Position", "Leave Type", "Date", "Days"]]

    for r in rows:
        table_data.append([
            r["full_name"],
            r["position"],
            r["leave_type"],
            f"{r['start_date']} → {r['end_date']}",
            r["total_days"]
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.darkblue),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("ALIGN", (0,0), (-1,-1), "CENTER")
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Team_Leave_Report_{department}.pdf"
    )
    
def normalize_leave_type(t):
    return (t or "").strip().upper()
    
@app.route("/leave-report/department/preview")
@admin_required
def preview_leave_report_department():
    from datetime import datetime, timedelta, date

    year = request.args.get("year", datetime.now().strftime("%Y"))
    month = request.args.get("month", datetime.now().strftime("%m")) 
    department = request.args.get("department", "all")

    conn = get_db()
    cur = conn.cursor()

    params = [year, month]
    dept_filter = ""

    first_day = f"{year}-{month}-01"
    last_day = (
        datetime.strptime(first_day, "%Y-%m-%d")
        + timedelta(days=32)
    ).replace(day=1) - timedelta(days=1)

    params = [last_day.strftime("%Y-%m-%d"), first_day]

    if department != "all":
        dept_filter = "AND d.name = ?"
        params.append(department)
    else:
        dept_filter = ""

    cur.execute(f"""
        SELECT
            u.id AS user_id,
            u.full_name,
            la.leave_type,
            la.start_date,
            la.end_date,
            u.entitlement,
            d.name AS department
        FROM leave_applications la
        JOIN users u ON u.id = la.user_id
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE la.status = 'Approved'
        AND la.start_date <= ?
        AND la.end_date >= ?
        {dept_filter}
        ORDER BY u.full_name
    """, params)

    rows = cur.fetchall()
    conn.close()

    # ✅ LEAVE TYPE → CELL CODE
    code_map = {
        "ANNUAL": "AL",
        "ANNUAL LEAVE": "AL",
        "NORMAL": "AL",
        "NORMAL LEAVE": "AL",
        "EMERGENCY": "AL",
        "EMERGENCY LEAVE": "AL",
        
        "COMPASSIONATE": "CL",
        "COMPASSIONATE LEAVE": "CL",
        "DEATH OF IMMEDIATE FAMILY MEMBERS": "CL",
        "DISASTER (FLOOD/FIRE)": "CL",

        "LEAVE-IN-LIEU": "LIL",
        "UNPAID": "UL",
        "UNPAID LEAVE": "UL",
        "MATERNITY": "MP",
        "PATERNITY": "MP",
        
        "MC": "MC"
    }

    users = {}

    for r in rows:
        uid = r["user_id"]

        if uid not in users:
            users[uid] = {
                "full_name": r["full_name"],
                "department": r["department"],
                "entitlement": r["entitlement"] or 0,
                "total_used": 0,
                "remaining": r["entitlement"] or 0,
                "daily": {}   
            }

        start = datetime.strptime(r["start_date"], "%Y-%m-%d")
        end   = datetime.strptime(r["end_date"], "%Y-%m-%d")

        cur_day = start
        while cur_day <= end:

            # ✅ ONLY MARK SELECTED MONTH-YEAR
            if cur_day.strftime("%Y") == year and cur_day.strftime("%m") == month:
                day_no = int(cur_day.strftime("%d"))
                leave_key = normalize_leave_type(r["leave_type"])
                users[uid]["daily"][day_no] = code_map.get(leave_key, "")

                # ✅ DEDUCT ONLY NON-MC & WEEKDAYS
                if r["leave_type"] != "MC" and cur_day.weekday() < 5:
                    users[uid]["total_used"] += 1
                    users[uid]["remaining"] -= 1

            cur_day += timedelta(days=1)

    return render_template(
        "reports/leave_report_department_preview.html",
        rows=list(users.values()),
        year=year,
        month=month,
        department=department,
        printed_date=date.today().strftime("%d-%m-%Y")
    )


@app.context_processor
def inject_current_year():
    from datetime import datetime
    return {
        "current_year": datetime.now().year
    }

if __name__ == "__main__":
    app.run()


# if __name__ == "__main__":
#     app.run(
#         host="0.0.0.0",
#         port=5000,
#         ssl_context=(
#             r"C:\Users\User\Desktop\Leave 1.0\localhost+1.pem",
#             r"C:\Users\User\Desktop\Leave 1.0\localhost+1-key.pem"
#         ),
#         debug=True
#     )

    
